import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinterdnd2 import DND_FILES, TkinterDnD

class CSVProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("CSV 面積計算工具")
        self.root.geometry("300x120")
        self.root.rowconfigure(4, weight=1)
        self.root.columnconfigure(1, weight=1)
        
        # 檔案路徑輸入框
        self.file_path_label = tk.Label(root, text="檔案路徑:")
        self.file_path_label.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        self.file_path_entry = tk.Entry(root, width=50)
        self.file_path_entry.grid(row=0, column=1, padx=10, pady=5, sticky=tk.EW)
        self.browse_button = tk.Button(root, text="瀏覽", command=self.browse_file)
        self.browse_button.grid(row=0, column=2, padx=10, pady=5)

        # 支持拖曳文件至檔案路徑輸入框
        self.file_path_entry.drop_target_register(DND_FILES)
        self.file_path_entry.dnd_bind('<<Drop>>', self.drop_file)
        
        # 數字選項 (小數點四捨五入)
        self.round_label = tk.Label(root, text="小數點四捨五入位數:")
        self.round_label.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        self.round_var = tk.IntVar(value=6)
        self.round_spinbox = tk.Spinbox(root, from_=0, to=10, textvariable=self.round_var, width=5)
        self.round_spinbox.grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)
        
        # 執行按鈕
        self.process_button = tk.Button(root, text="執行", command=self.process_csv)
        self.process_button.grid(row=2, column=1, pady=5, sticky=tk.W)

    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if file_path:
            self.file_path_entry.delete(0, tk.END)
            self.file_path_entry.insert(0, file_path)

    def drop_file(self, event):
        file_path = event.data.strip('{}')
        if file_path:
            self.file_path_entry.delete(0, tk.END)
            self.file_path_entry.insert(0, file_path)

    def process_csv(self):
        file_path = self.file_path_entry.get()
        if not file_path:
            messagebox.showerror("錯誤", "請選擇檔案路徑")
            return
        
        # 更改輸出的 Excel 檔案名稱與輸入的 CSV 檔案名稱相同，只更改副檔名
        output_path = file_path.replace('.csv', '.xlsx')
        
        round_digits = self.round_var.get()
        
        try:
            # 讀取 CSV 文件

            try:
                df = pd.read_csv(file_path, encoding='utf-8')
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, encoding='ISO-8859-1')

            #df = pd.read_csv(file_path)
            
            # 假設欄位名稱分別是 "圖層" 和 "面積"
            df['面積'] = df['面積'].round(round_digits)
            grouped = df.groupby('圖層')['面積'].sum().round(round_digits)
            
            # 創建一個新的 DataFrame 存儲結果
            output_data = []
            for layer, total_area in grouped.items():
                # 獲取該圖層所有對應的面積值
                areas = df[df['圖層'] == layer]['面積'].tolist()
                # 添加每一筆資料
                areasm = list(map(lambda x: round(x / 10000, round_digits), areas))
                for area in areasm:
                    output_data.append([layer, area])
                # 添加面積加總和計算公式
                area_str = '+'.join([f"{area:.{round_digits}f}" for area in areasm if area])
                output_data.append(["數值總數", len(areasm)])
                output_data.append(["面積加總", sum(areasm)])
                output_data.append(["計算公式", area_str])
                
            
            # 將結果存儲到新的 DataFrame
            result_df = pd.DataFrame(output_data, columns=['圖層', '面積'])
            
            # 先將結果寫入 Excel 文件
            result_df.to_excel(output_path, index=False, engine='openpyxl')
            
            # 使用 openpyxl 打開 Excel 檔案並設定格式
            wb = load_workbook(output_path)
            ws = wb.active

            # 設置格式：保留 n 位小數
            format_string = f"0.{''.join(['0' for _ in range(round_digits)])}"
            for row in ws.iter_rows(min_row=2, max_row=len(result_df) + 1, min_col=2, max_col=2):
                for cell in row:
                    # 設置數值格式，保留指定的小數位數
                    cell.number_format = format_string

            # 保存修改後的 Excel 檔案
            wb.save(output_path)
            
            messagebox.showinfo("完成", f"結果已寫入 {output_path}")
        except Exception as e:
            messagebox.showerror("錯誤", str(e))

if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = CSVProcessorApp(root)
    root.mainloop()
